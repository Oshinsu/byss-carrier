import sys
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
import json
import re
from pathlib import Path

BASE = 'C:/Users/Gary/Desktop/BYSS GROUP/BYSS incroyable'
OUT = 'C:/Users/Gary/Desktop/BYSS GROUP/byss-carrier/supabase/migrations/007_import_all_xlsx.sql'

def esc(val):
    """Escape a string for SQL single-quote literals."""
    if val is None:
        return 'NULL'
    s = str(val).strip()
    if not s:
        return 'NULL'
    return "'" + s.replace("'", "''") + "'"

SECTOR_DOMAIN = {
    'Télécoms': 'economique',
    'Grande Distribution': 'economique',
    'Banque': 'economique',
    'Assurance': 'economique',
    'Automobile': 'economique',
    'Automobile/Distribution': 'economique',
    'Agroalimentaire': 'economique',
    'BTP': 'economique',
    'Immobilier': 'economique',
    'Industrie': 'economique',
    'Transport': 'economique',
    'Transport/Nautisme': 'economique',
    'Location véhicules': 'economique',
    'Énergie': 'economique',
    'Imprimerie': 'economique',
    'Sécurité Sociale': 'institutionnelle',
    'Institution': 'institutionnelle',
    'Patronat': 'institutionnelle',
    'Patronat Outre-Mer': 'institutionnelle',
    'Formation': 'institutionnelle',
    'Défense/Géopolitique': 'institutionnelle',
    'Politique': 'politique',
    'Média': 'media',
    'Média/Réseau': 'media',
    'Production': 'media',
    'Agence Com': 'media',
    'Hôtellerie': 'economique',
    'Tourisme': 'economique',
    'Tourisme/Événementiel': 'economique',
    'Restaurant': 'economique',
    'Restaurant/Bar/Club': 'economique',
    'Restaurant/Cave': 'economique',
    'Bar cocktails': 'economique',
    'Bar/Club': 'economique',
    'Bar/Live music': 'economique',
    'Bar/Lounge': 'economique',
    'Bar/Rooftop': 'economique',
    'Bar/Restaurant': 'economique',
    'Aérien': 'economique',
    'Comptable': 'economique',
    'Conseil': 'economique',
    'Entreprise': 'economique',
    'Environnement': 'sociale',
    'RH/Recrutement': 'economique',
    'Réseau': 'sociale',
}

lines = []
lines.append("-- ═══════════════════════════════════════════════════════════════")
lines.append("-- 007_import_all_xlsx.sql")
lines.append("-- Full XLSX data import: contacts, bible de vente, grille tarifaire, templates")
lines.append("-- Generated from BYSS incroyable Excel files")
lines.append("-- ═══════════════════════════════════════════════════════════════")
lines.append("")
lines.append("BEGIN;")
lines.append("")

# ═══════════════════════════════════════════════════════
# 1. ALTER lore_entries to accept 'bible' universe
# ═══════════════════════════════════════════════════════
lines.append("-- Extend universe CHECK to include 'bible'")
lines.append("ALTER TABLE lore_entries DROP CONSTRAINT IF EXISTS lore_entries_universe_check;")
lines.append("ALTER TABLE lore_entries ADD CONSTRAINT lore_entries_universe_check")
lines.append("  CHECK (universe IN ('cadifor','jurassic_wars','eveil','toxic','lignee','bible'));")
lines.append("")

# ═══════════════════════════════════════════════════════
# 2. CONTACTS -> intel_entities (291 rows)
# ═══════════════════════════════════════════════════════
lines.append("-- ═══════════════════════════════════════════════════════════════")
lines.append("-- INTEL_ENTITIES — contacts from contacts_martinique_byss_group_ULTRA_FINAL.xlsx")
lines.append("-- ═══════════════════════════════════════════════════════════════")
lines.append("")

wb = openpyxl.load_workbook(f'{BASE}/contacts_martinique_byss_group_ULTRA_FINAL.xlsx', read_only=True, data_only=True)
ws = wb['Contacts Martinique']

contact_count = 0
for row in ws.iter_rows(min_row=2, max_col=9, values_only=True):
    nom = row[0]
    poste = row[1]
    entreprise = row[2]
    secteur = row[3]
    telephone = row[4]
    email = row[5]
    linkedin = row[6]
    source = row[7]
    notes = row[8]

    # Skip empty rows and section headers (rows where only Nom is filled)
    if not nom:
        continue
    if not secteur and not poste and not entreprise:
        continue

    domain = SECTOR_DOMAIN.get(secteur, 'economique')

    # Build contacts JSONB
    contact_info = {}
    if poste:
        contact_info['poste'] = str(poste).strip()
    if telephone:
        contact_info['telephone'] = str(telephone).strip()
    if email:
        contact_info['email'] = str(email).strip()
    if linkedin:
        contact_info['linkedin'] = str(linkedin).strip()
    if source:
        contact_info['source'] = str(source).strip()

    contacts_json = json.dumps([contact_info], ensure_ascii=False)

    # Build tags
    tags = []
    if secteur:
        tags.append(str(secteur).strip())
    if source:
        tags.append(str(source).strip())

    tags_sql = "ARRAY[" + ",".join(esc(t) for t in tags) + "]" if tags else "'{}'::text[]"

    entity_name = str(nom).strip()
    entity_type = str(secteur).strip() if secteur else 'Entreprise'
    description = ''
    if entreprise:
        description += str(entreprise).strip()
    if poste:
        description += ' -- ' + str(poste).strip()
    if not description:
        description = entity_name

    notes_sql = esc(notes) if notes else 'NULL'

    lines.append(f"INSERT INTO intel_entities (domain, name, type, description, contacts, notes, tags)")
    lines.append(f"VALUES ({esc(domain)}, {esc(entity_name)}, {esc(entity_type)}, {esc(description)}, {esc(contacts_json)}::jsonb, {notes_sql}, {tags_sql})")
    lines.append(f"ON CONFLICT DO NOTHING;")
    lines.append("")
    contact_count += 1

wb.close()
lines.append(f"-- Total contacts inserted: {contact_count}")
lines.append("")

# ═══════════════════════════════════════════════════════
# 3. BIBLE DE VENTE -> lore_entries
# ═══════════════════════════════════════════════════════
lines.append("-- ═══════════════════════════════════════════════════════════════")
lines.append("-- LORE_ENTRIES — Bible de Vente")
lines.append("-- ═══════════════════════════════════════════════════════════════")
lines.append("")

wb = openpyxl.load_workbook(f'{BASE}/bible_vente_byss_group.xlsx', read_only=True, data_only=True)
ws = wb['Document']

current_chapter = None
chapter_idx = 0
entry_idx = 0
bible_count = 0

for row in ws.iter_rows(min_row=1, max_col=2, values_only=True):
    col_a = row[0]
    col_b = row[1]

    if not col_a and not col_b:
        continue

    col_a_str = str(col_a).strip() if col_a else ''
    col_b_str = str(col_b).strip() if col_b else ''

    # Detect chapter headers
    if col_a_str.startswith('CHAPITRE') or col_a_str.startswith('BIBLE DE VENTE'):
        current_chapter = col_a_str
        chapter_idx += 1
        entry_idx = 0
        title = col_a_str
        content = col_b_str if col_b_str else ''
        category = 'chapter'
        word_count = len(content.split()) if content else 0

        lines.append(f"INSERT INTO lore_entries (universe, title, content, category, tags, word_count, order_index)")
        lines.append(f"VALUES ('bible', {esc(title)}, {esc(content)}, 'chapter', ARRAY['bible','vente'], {word_count}, {chapter_idx * 100})")
        lines.append(f"ON CONFLICT DO NOTHING;")
        lines.append("")
        bible_count += 1
        continue

    # Regular entries
    entry_idx += 1
    title = col_a_str if col_a_str else (current_chapter or 'Bible de Vente')
    content = col_b_str if col_b_str else col_a_str

    if title == content and not col_b_str:
        category = 'section'
    else:
        category = 'entry'

    word_count = len(content.split()) if content else 0

    lines.append(f"INSERT INTO lore_entries (universe, title, content, category, tags, word_count, order_index)")
    lines.append(f"VALUES ('bible', {esc(title)}, {esc(content)}, {esc(category)}, ARRAY['bible','vente'], {word_count}, {chapter_idx * 100 + entry_idx})")
    lines.append(f"ON CONFLICT DO NOTHING;")
    lines.append("")
    bible_count += 1

wb.close()
lines.append(f"-- Total bible entries inserted: {bible_count}")
lines.append("")

# ═══════════════════════════════════════════════════════
# 4. GRILLE TARIFAIRE -> lore_entries (universe='bible', category='pricing')
# ═══════════════════════════════════════════════════════
lines.append("-- ═══════════════════════════════════════════════════════════════")
lines.append("-- LORE_ENTRIES — Grille Tarifaire (pricing)")
lines.append("-- ═══════════════════════════════════════════════════════════════")
lines.append("")

wb = openpyxl.load_workbook(f'{BASE}/grille_tarifaire_byss_group.xlsx', read_only=True, data_only=True)
ws = wb['Grille Tarifaire']

tarif_count = 0
current_service = None
order = 2000

for row in ws.iter_rows(min_row=1, max_col=4, values_only=True):
    vals = [str(v).strip() if v else '' for v in row]

    if not any(vals):
        continue

    # Detect service headers
    if 'SERVICE' in vals[0].upper() and ('\u2014' in vals[0] or '-' in vals[0]):
        current_service = vals[0]
        order += 1
        lines.append(f"INSERT INTO lore_entries (universe, title, content, category, tags, word_count, order_index)")
        lines.append(f"VALUES ('bible', {esc(current_service)}, {esc(current_service)}, 'pricing_section', ARRAY['bible','tarif','pricing'], 0, {order})")
        lines.append(f"ON CONFLICT DO NOTHING;")
        lines.append("")
        tarif_count += 1
        continue

    # Skip column headers and title rows
    if vals[0] == 'SERVICE' or 'GRILLE TARIFAIRE OFFICIELLE' in vals[0].upper() or vals[0].startswith('Tarifs HT'):
        continue

    # Data rows
    if vals[0] and vals[1]:
        order += 1
        service_name = vals[0]
        prix = vals[1]
        recurrence = vals[2]
        detail = vals[3]

        content_parts = [f"Prix: {prix}"]
        if recurrence:
            content_parts.append(f"R\u00e9currence: {recurrence}")
        if detail:
            content_parts.append(f"D\u00e9tail: {detail}")
        content = ' | '.join(content_parts)
        word_count = len(content.split())

        lines.append(f"INSERT INTO lore_entries (universe, title, content, category, tags, word_count, order_index)")
        lines.append(f"VALUES ('bible', {esc(service_name)}, {esc(content)}, 'pricing', ARRAY['bible','tarif','pricing'], {word_count}, {order})")
        lines.append(f"ON CONFLICT DO NOTHING;")
        lines.append("")
        tarif_count += 1

wb.close()
lines.append(f"-- Total pricing entries inserted: {tarif_count}")
lines.append("")

# ═══════════════════════════════════════════════════════
# 5. TEMPLATES -> prompts table
# ═══════════════════════════════════════════════════════
lines.append("-- ═══════════════════════════════════════════════════════════════")
lines.append("-- PROMPTS — Templates de prospection")
lines.append("-- ═══════════════════════════════════════════════════════════════")
lines.append("")

wb = openpyxl.load_workbook(f'{BASE}/templates_prospection_byss_group.xlsx', read_only=True, data_only=True)

template_count = 0
channel_map = {
    'Emails Prospection': 'email',
    'Scripts T\u00e9l\u00e9phone': 'telephone',
    'WhatsApp & SMS': 'whatsapp',
    'DM Instagram': 'instagram_dm',
}

def flush_template(channel, section, parts):
    """Generate an INSERT for a template."""
    if not parts or not section:
        return 0
    template_text = '\n\n'.join(f"[{k}]\n{v}" for k, v in parts.items())
    variables = []
    for match in re.findall(r'\[([A-Z_]+)\]', template_text):
        if match not in variables:
            variables.append(match)
    vars_json = json.dumps(variables, ensure_ascii=False)

    name = f"{channel} -- {section}"
    lines.append(f"INSERT INTO prompts (name, category, template, variables, model, is_master)")
    lines.append(f"VALUES ({esc(name)}, 'text', {esc(template_text)}, {esc(vars_json)}::jsonb, 'prospection', false)")
    lines.append(f"ON CONFLICT DO NOTHING;")
    lines.append("")
    return 1

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    channel = channel_map.get(sheet_name, 'text')

    current_section = None
    template_parts = {}

    for row in ws.iter_rows(min_row=1, max_col=2, values_only=True):
        col_a = str(row[0]).strip() if row[0] else ''
        col_b = str(row[1]).strip() if row[1] else ''

        if not col_a and not col_b:
            template_count += flush_template(channel, current_section, template_parts)
            template_parts = {}
            continue

        # Section headers (emoji lines)
        if any(e in col_a for e in ['\U0001f4e7', '\U0001f4de', '\U0001f4ac', '\U0001f4f1']) and '\u2014' in col_a:
            template_count += flush_template(channel, current_section, template_parts)
            template_parts = {}
            current_section = col_a
            continue

        if col_a and col_b:
            template_parts[col_a] = col_b
        elif col_a and not col_b:
            if current_section:
                template_parts['Note'] = col_a

    # Flush last template in sheet
    template_count += flush_template(channel, current_section, template_parts)

wb.close()
lines.append(f"-- Total templates inserted: {template_count}")
lines.append("")

lines.append("COMMIT;")
lines.append("")
lines.append(f"-- ═══════════════════════════════════════════════════════════════")
lines.append(f"-- SUMMARY:")
lines.append(f"--   intel_entities: {contact_count} contacts")
lines.append(f"--   lore_entries (bible): {bible_count} entries")
lines.append(f"--   lore_entries (pricing): {tarif_count} entries")
lines.append(f"--   prompts (templates): {template_count} templates")
lines.append(f"--   TOTAL: {contact_count + bible_count + tarif_count + template_count} rows")
lines.append(f"-- ═══════════════════════════════════════════════════════════════")

# Write output
output = '\n'.join(lines)
Path(OUT).parent.mkdir(parents=True, exist_ok=True)
with open(OUT, 'w', encoding='utf-8') as f:
    f.write(output)

print(f"Written {len(lines)} lines to {OUT}")
print(f"  intel_entities: {contact_count} contacts")
print(f"  lore_entries (bible): {bible_count} entries")
print(f"  lore_entries (pricing): {tarif_count} entries")
print(f"  prompts (templates): {template_count} templates")
print(f"  TOTAL: {contact_count + bible_count + tarif_count + template_count} rows")
